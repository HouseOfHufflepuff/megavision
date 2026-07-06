"""
Repeatable roster + financials sync for the 13 MEGAVISION team pages.

This script holds NO copy of the spreadsheet and never caches one on disk.
Every run takes a freshly-fetched export as input, uses it in place, and the
caller deletes it immediately after. There is no standalone Google credential
here, so the live fetch itself is done by Claude via the Google Drive
connector -- the full cycle, every time, is:

    1. Claude fetches the live spreadsheet export to a throwaway temp path
    2. python3 update_rosters.py --xlsx /path/to/that/temp/file [--push]
    3. the temp file is deleted

Never point --xlsx at a path inside this repo, and never commit the xlsx.
"""
import argparse
import subprocess
import sys
from pathlib import Path

import openpyxl

from common import TEAMS, head, FOOT

parser = argparse.ArgumentParser()
parser.add_argument("--xlsx", required=True, help="Path to a freshly-fetched spreadsheet export (temp file, not stored in this repo)")
parser.add_argument("--push", action="store_true")
args = parser.parse_args()

SS_PATH = Path(args.xlsx)
if not SS_PATH.exists():
    print(f"ERROR: {SS_PATH} not found.", file=sys.stderr)
    sys.exit(1)


def parse_team_tab(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
    stadium_name = rows[0][1] or ""
    capacity = rows[0][6]

    header_idx = None
    for i, r in enumerate(rows):
        if r[1] == "Player" and r[2] == "Pos":
            header_idx = i
            break
    if header_idx is None:
        return None

    year_labels = [rows[header_idx][3], rows[header_idx][4], rows[header_idx][5]]

    roster = []
    for r in rows[header_idx + 1:]:
        if r[0] == "Total":
            break
        if r[1] is None:
            continue
        roster.append({
            "player": r[1],
            "pos": r[2] or "",
            "y1": r[3],
            "y2": r[4],
            "y3": r[5],
            "buyout": r[6],
        })

    return {
        "stadium": stadium_name,
        "capacity": capacity,
        "year_labels": year_labels,
        "roster": roster,
    }


def money(v):
    if v is None:
        return "—"
    if isinstance(v, (int, float)):
        return f"${v:,.2f}"
    return str(v)


wb = openpyxl.load_workbook(str(SS_PATH), data_only=True)

updated = []
for code, name, owners in TEAMS:
    if code not in wb.sheetnames:
        print(f"WARN: no tab for {code}, skipping", file=sys.stderr)
        continue
    data = parse_team_tab(wb[code])
    if data is None:
        print(f"WARN: could not find roster header for {code}, skipping", file=sys.stderr)
        continue

    roster = data["roster"]
    roster_size = len(roster)
    total_payroll = sum(p["y1"] for p in roster if isinstance(p["y1"], (int, float)))
    pos_counts = {}
    for p in roster:
        pos_counts[p["pos"]] = pos_counts.get(p["pos"], 0) + 1
    pos_summary = " &middot; ".join(f"{v} {k}" for k, v in sorted(pos_counts.items()) if k)
    season_net = -total_payroll  # no games/revenue yet this season

    y1_label, y2_label, y3_label = data["year_labels"]
    cap = data["capacity"]
    capacity_str = f"{cap:,.0f}" if isinstance(cap, (int, float)) else str(cap or "—")
    roster_rows = []
    def sort_key(p):
        return -p["y1"] if isinstance(p["y1"], (int, float)) else 0

    for p in sorted(roster, key=sort_key):
        roster_rows.append(
            f'<tr><td>{p["player"]}</td><td>{p["pos"]}</td>'
            f'<td>{money(p["y1"])}</td><td>{money(p["y2"])}</td><td>{money(p["y3"])}</td>'
            f'<td class="dim">{money(p["buyout"])}</td></tr>'
        )

    slug = code.lower()
    page = head(name, "teams.html") + f"""
    <div class="mv-page-header">
      <h1 class="mv-chrome-text">{name}<span class="mv-badge">{code}</span></h1>
      <div class="sub">{", ".join(owners)} &middot; {data["stadium"]} (Capacity {capacity_str})</div>
    </div>

    <div class="mv-stat-grid">
      <div class="mv-stat"><div class="label">Record</div><div class="value">0-0-0</div></div>
      <div class="mv-stat"><div class="label">Points</div><div class="value">0</div></div>
      <div class="mv-stat"><div class="label">League Rank</div><div class="value">&mdash;</div></div>
      <div class="mv-stat"><div class="label">Roster Size</div><div class="value">{roster_size}</div></div>
      <div class="mv-stat"><div class="label">Total Payroll</div><div class="value">{money(total_payroll)}</div></div>
      <div class="mv-stat"><div class="label">Season Net</div><div class="value">{money(season_net)}</div></div>
    </div>

    <section class="card mv-card">
      <h2 class="mv-chrome-text">Roster</h2>
      <div class="sub">{roster_size} players &middot; {pos_summary} &middot; no games played yet this season</div>
      <div class="mv-table-scroll">
        <table class="mv-table">
          <thead><tr><th>Player</th><th>Pos</th><th>{y1_label}</th><th>{y2_label}</th><th>{y3_label}</th><th>BuyOut</th></tr></thead>
          <tbody>
            {"".join(roster_rows)}
          </tbody>
        </table>
      </div>
    </section>

    <p style="margin-top:24px;"><a href="teams.html" style="color:var(--mv-ink-muted);font-size:13px;">&larr; Back to all teams</a></p>
""" + FOOT

    with open(f"team-{slug}.html", "w") as f:
        f.write(page)
    updated.append((code, roster_size, total_payroll))

print(f"Updated {len(updated)} team pages:")
for code, size, payroll in updated:
    print(f"  {code}: {size} players, ${payroll:,.2f} payroll")

if args.push:
    files = [f"team-{c.lower()}.html" for c, _, _ in updated]
    subprocess.run(["git", "add"] + files, check=True)
    result = subprocess.run(["git", "diff", "--cached", "--quiet"])
    if result.returncode == 0:
        print("No changes to publish.")
    else:
        subprocess.run(
            ["git", "-c", "user.email=ahrens@gmail.com", "-c", "user.name=Jeremy Ahrens",
             "commit", "-q", "-m", "Update team rosters and financials from spreadsheet"],
            check=True,
        )
        subprocess.run(["git", "push", "-q"], check=True)
        print("Pushed to GitHub Pages.")
