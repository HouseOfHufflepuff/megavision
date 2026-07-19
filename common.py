# Shared nav/head/foot template + team roster used by build_site.py and update_rosters.py

import io
import urllib.request
from datetime import datetime, timezone

import openpyxl

SHEET_ID = "1l-BG5on67L9beLArTMShJiVZgJOL3CUed3rbuwZKBL0"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"


def fetch_live_workbook():
    """Always live: the sheet is shared 'anyone with the link can view', so
    this is a plain unauthenticated HTTP GET, straight into memory. Never
    write the export to disk, never cache it between runs."""
    req = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def find_team_sheet(wb, code):
    """Team tabs occasionally get renamed in the live sheet (e.g. 'FAV' ->
    'FAV-Updated'). Try an exact match first, then fall back to any sheet
    name starting with the code, instead of silently dropping the team."""
    if code in wb.sheetnames:
        return code
    for name in wb.sheetnames:
        if name.upper().startswith(code.upper()):
            return name
    return None


def fetch_stadiums(wb):
    """code -> {stadium, capacity}, from row 1 of each team tab."""
    stadiums = {}
    for code, _, _ in TEAMS:
        sheet_name = find_team_sheet(wb, code)
        if sheet_name is None:
            continue
        row0 = next(wb[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))
        stadiums[code] = {"stadium": row0[1] or "", "capacity": row0[6]}
    return stadiums


def fetch_season_salary_totals(wb, label):
    """code -> team payroll total for the given season label (e.g. '25/26'),
    read straight from each team tab's own 'Total' row so it matches
    whatever the sheet itself computes. Returns None for a team if that
    season label isn't one of its three year columns."""
    totals = {}
    for code, _, _ in TEAMS:
        sheet_name = find_team_sheet(wb, code)
        if sheet_name is None:
            continue
        rows = list(wb[sheet_name].iter_rows(min_row=1, max_row=40, values_only=True))
        header_idx = next((i for i, r in enumerate(rows) if r[1] == "Player" and r[2] == "Pos"), None)
        if header_idx is None:
            continue
        year_labels = [rows[header_idx][3], rows[header_idx][4], rows[header_idx][5]]
        if label not in year_labels:
            totals[code] = None
            continue
        col = 3 + year_labels.index(label)
        total_row = next((r for r in rows[header_idx + 1:] if r[0] == "Total"), None)
        totals[code] = total_row[col] if total_row else None
    return totals


def fetch_all_season_labels(wb):
    """Every distinct season label (e.g. '25/26') appearing as a year column
    on any team tab, sorted oldest-first. Used to populate the financials
    page's season dropdown -- whatever's actually in the sheet, live."""
    labels = set()
    for code, _, _ in TEAMS:
        sheet_name = find_team_sheet(wb, code)
        if sheet_name is None:
            continue
        rows = list(wb[sheet_name].iter_rows(min_row=1, max_row=10, values_only=True))
        header_idx = next((i for i, r in enumerate(rows) if r[1] == "Player" and r[2] == "Pos"), None)
        if header_idx is None:
            continue
        for v in (rows[header_idx][3], rows[header_idx][4], rows[header_idx][5]):
            if v:
                labels.add(v)
    return sorted(labels, key=lambda s: tuple(int(p) for p in s.split("/")))


def fetch_league_schedule_games(wb, weeks=range(1, 23)):
    """Every fixture row from the 'League Schedule' tab for the given weeks
    (default: the 22 regular-season weeks), with team names resolved to our
    codes. Each entry: week, home/away code, stadium, capacity, fan
    interest (home/away/total), attendance, ticket price, gate receipts,
    home/away revenue -- straight off the sheet's own computed values."""
    ws = wb["League Schedule"]
    games = []
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        week = r[0]
        if week is None or week not in weeks:
            continue
        home_code = resolve_team_code(r[2])
        away_code = resolve_team_code(r[3])
        if not home_code or not away_code:
            continue
        games.append({
            "week": int(week),
            "home": home_code, "away": away_code,
            "stadium": r[4], "capacity": r[5],
            "home_int": r[6], "away_int": r[7], "total_int": r[8],
            "attendance": r[9], "ticket_price": r[10], "gate_receipts": r[11],
            "home_rev": r[12], "away_rev": r[13],
        })
    return games


def fetch_standings_reference(wb):
    """The two static reference tables the sheet's own Fan Formula reads
    from, straight off the 'Standings' tab:
      - rank_bonus: {rank int -> Standings fan bonus} (Standings!A3:A14 / M3:M14)
      - points_bonus: {team code -> fan bonus} for the top-5 teams by
        points (Standings!F16:G20)
    """
    ws = wb["Standings"]
    rows = list(ws.iter_rows(min_row=1, max_row=25, values_only=True))
    rank_bonus = {}
    for r in rows[2:14]:  # rows 3-14
        if isinstance(r[0], (int, float)) and isinstance(r[12], (int, float)):
            rank_bonus[int(r[0])] = r[12]
    points_bonus = {}
    for r in rows[15:20]:  # rows 16-20
        code, bonus = r[5], r[6]
        if code and isinstance(bonus, (int, float)):
            points_bonus[code] = bonus
    return rank_bonus, points_bonus


def fetch_firm_legacy(wb):
    """code -> Firm + Legacy fan total, from the 'Firms and Legacy' tab's
    own 'Firm + Legacy' column (U)."""
    ws = wb["Firms and Legacy"]
    out = {}
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = resolve_team_code(r[0])
        if code and isinstance(r[20], (int, float)):
            out[code] = r[20]
    return out


def compute_fan_formula(rank_bonus, points_bonus, firm_legacy, fantrax_standings, top_xi, mbp):
    """The full live Fan Formula for every team with Fantrax standings data:
    Standings (live rank tier) + Firm+Legacy (sheet) + Top XI (20/player,
    live from Fantrax) + MBP (50 if you own the top scorer, live from
    Fantrax) + Points Bonus (sheet, static top-5 table). MS8 has no active
    Fantrax roster this season and is intentionally excluded.
    Returns code -> {standings, firm_legacy, top_xi, mbp, points_bonus, total, xi_count}.
    """
    xi_counts = {}
    for p in top_xi:
        xi_counts[p["code"]] = xi_counts.get(p["code"], 0) + 1

    out = {}
    for code, s in fantrax_standings.items():
        standings_v = rank_bonus.get(s["rank"], 0)
        legacy_v = firm_legacy.get(code, 0)
        xi_count = xi_counts.get(code, 0)
        xi_v = xi_count * 20
        mbp_v = 50 if mbp and mbp["code"] == code else 0
        pts_v = points_bonus.get(code, 0)
        out[code] = {
            "standings": standings_v, "firm_legacy": legacy_v,
            "top_xi": xi_v, "xi_count": xi_count, "mbp": mbp_v,
            "points_bonus": pts_v,
            "total": standings_v + legacy_v + xi_v + mbp_v + pts_v,
        }
    return out


def fetch_youth(wb):
    """code -> list of youth-drafted players (all-time), most recent first,
    straight from the 'Youth' tab. Columns: Year, Team, Player, Pos,
    Age, Current Owner Team, Currently Playing, Original Team,
    Fantrax Eligible, Promoted, Youth Rights Released, Contract Frozen."""
    ws = wb["Youth"]
    rows = list(ws.iter_rows(min_row=2, max_row=1167, values_only=True))
    valid_codes = {code for code, _, _ in TEAMS}
    youth_by_code = {code: [] for code in valid_codes}
    for row in rows:
        if not row or row[1] not in valid_codes or not row[2]:
            continue
        code = row[1]
        promoted = str(row[9] or "").strip().upper() == "Y"
        released = str(row[10] or "").strip().upper() == "Y"
        frozen = str(row[11] or "").strip().upper() == "Y"
        if promoted:
            status = "Promoted"
        elif released:
            status = "Released"
        elif frozen:
            status = "Frozen"
        else:
            status = "Active"
        youth_by_code[code].append({
            "year": row[0],
            "player": row[2],
            "pos": row[3] or "",
            "age": row[4],
            "club": row[6] or row[5] or row[7] or "",
            "status": status,
        })
    return youth_by_code


def fetch_trophy_room(wb):
    ws = wb["Trophy Room"]
    all_rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))
    header_idx = None
    for i, row in enumerate(all_rows):
        if row and row[1] == "Premium Title":
            header_idx = i
            break
    comps = [c for c in all_rows[header_idx][1:8] if c]
    seasons = []
    for row in all_rows[header_idx + 1:]:
        if row[0] == "Money":
            break
        if row[0]:
            seasons.append(list(row))
    return comps, seasons

# code, full name, owners
TEAMS = [
    ("FAV", "5th Ave Argyle", ["Jack Weatherman"]),
    ("POW", "Battersea Power Bottoms", ["Sam Rufer", "Joe Effertz"]),
    ("CRG", "CRG McGovern", ["Casey McGovern", "Ryan McGovern", "Grady McGovern", "Jadyn McGovern"]),
    ("DU", "Divided United", ["Chris Gauron"]),
    ("HUF", "House of Hufflepuff", ["Jeremy Ahrens"]),
    ("MS8", "MS 08th", ["Raul Templonuevo"]),
    ("NAC", "NFC Andover City", ["Kris Lien"]),
    ("QFC", "Quidpool FC", ["Erik Johnson"]),
    ("REN", "Real News", ["Reid Foster"]),
    ("BHB", "The Bookhouse Boys", ["Matt O'Laughlin"]),
    ("TTS", "Thottenham Thotspur", ["Kevin O'Laughlin"]),
    ("WTF", "What The FC", ["Erik Olson"]),
    ("ASS", "Wholeassed United FC", ["Kirk Walton", "Dan Hinrichs"]),
]

# Historical Trophy Room winner names don't always match a team's current
# full name exactly (renames, abbreviations) -- map every variant seen to a code.
TEAM_ALIASES = {
    "5th Ave Argyle": "FAV",
    "5th Avenue Argyle": "FAV",
    "Battersea Power Bottoms": "POW",
    "Battersea Power": "POW",
    "CRG McGovern": "CRG",
    "Divided United": "DU",
    "House of Hufflepuff": "HUF",
    "MS 08th": "MS8",
    "MS 08th FC": "MS8",
    "MS08": "MS8",
    "NFC Andover City": "NAC",
    "Andover City": "NAC",
    "Quidpool FC": "QFC",
    "Real News": "REN",
    "The Bookhouse Boys": "BHB",
    "Bookhouse Boys": "BHB",
    "Thottenham Thotspur": "TTS",
    "What The FC": "WTF",
    "What the FC": "WTF",
    "Wholeassed United FC": "ASS",
}


def resolve_team_code(name):
    if not name:
        return None
    return TEAM_ALIASES.get(name.strip())


def tally_trophies(comps, seasons):
    """code -> {comp: count}, built from Trophy Room season rows."""
    tally = {code: {c: 0 for c in comps} for code, _, _ in TEAMS}
    for row in seasons:
        for i, comp in enumerate(comps):
            winner = row[i + 1]
            code = resolve_team_code(winner)
            if code:
                tally[code][comp] += 1
    return tally


# Real EPL club abbreviations (as used in the sheet's messy "Player" field,
# already resolved by player_clean.clean_player) -> full club name.
CLUB_NAMES = {
    "ARS": "Arsenal", "AVL": "Aston Villa", "BHA": "Brighton", "BOU": "Bournemouth",
    "BRE": "Brentford", "BRF": "Brentford", "BRI": "Brighton", "BUR": "Burnley",
    "CHE": "Chelsea", "CPY": "Crystal Palace", "CRY": "Crystal Palace",
    "EVE": "Everton", "FUL": "Fulham", "LEE": "Leeds United", "LEEDS": "Leeds United",
    "LIV": "Liverpool", "MCI": "Manchester City", "MUN": "Manchester United",
    "NEW": "Newcastle United", "NOT": "Nottingham Forest", "SUN": "Sunderland",
    "TOT": "Tottenham Hotspur", "WHU": "West Ham United", "WOL": "Wolverhampton Wanderers",
}


def club_full_name(abbr):
    """Real club abbreviation/name -> full name. Passes through anything
    not in the map (already-full names, or a genuine data artifact in the
    sheet) rather than guessing or hiding it."""
    if not abbr:
        return None
    return CLUB_NAMES.get(abbr.upper(), abbr)


# Short labels so trophy tiles never wrap to 3 lines
COMP_ABBR = {
    "Premium Title": "Premium",
    "Champions League": "Champions Lg",
    "Europa League": "Europa Lg",
    "Mega FA Cup": "FA Cup",
    "Citadel Cup": "Citadel Cup",
    "Mega Community Shield": "Comm. Shield",
    "Mega Super Cup": "Super Cup",
}


def fetch_fans(wb):
    """code -> current fan count, from the Standings tab (not every team
    always has a row there -- missing means 0/unknown, shown as '—')."""
    ws = wb["Standings"]
    rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[1] == "Team":
            header_idx = i
            break
    fans = {}
    if header_idx is None:
        return fans
    for row in rows[header_idx + 1:]:
        if not row or not row[1]:
            break
        code = resolve_team_code(row[1])
        if code:
            fans[code] = row[12]
    return fans


def owner_short(owners):
    """'Jeremy Ahrens' -> 'Jeremy A.'; extra owners noted as '+N'."""
    first = owners[0].strip()
    parts = first.split()
    if len(parts) >= 2:
        short = f"{parts[0]} {parts[-1][0]}."
    else:
        short = first
    if len(owners) > 1:
        short += f" +{len(owners) - 1}"
    return short


POSITION_ORDER = ["GK", "D", "M", "F"]


def position_sort_key(pos):
    pos = (pos or "").upper()
    try:
        return POSITION_ORDER.index(pos)
    except ValueError:
        return len(POSITION_ORDER)


NAV_LINKS = [
    ("index.html", "Home"),
    ("teams.html", "Teams"),
    ("financials.html", "Financials"),
    ("snowmobile-lifestlye.html", "Snow"),
    ("armstrong-1.html", "Draft"),
    ("soccer.html", "Soccer"),
    ("pigeons/index.html", "Pigeons"),
    ("lazy-river/index.html", "Lazy River"),
]


def head(title, active):
    """No logo in the nav anywhere -- it's a big hero image at the top of
    every page's content instead (see hero_logo())."""
    nav_items = []
    for href, label in NAV_LINKS:
        cls = ' class="active"' if href == active else ""
        nav_items.append(f'<a href="{href}"{cls}>{label}</a>')
    nav_links = "\n      ".join(nav_items)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} — MEGAVISION</title>
<link rel="icon" type="image/png" sizes="32x32" href="favicon-32.png">
<link rel="icon" type="image/png" sizes="16x16" href="favicon-16.png">
<link rel="apple-touch-icon" href="apple-touch-icon.png">
<link rel="stylesheet" href="palette.css">
<link rel="stylesheet" href="dashboard.css">
</head>
<body>
  <nav class="mv-nav">
    <div class="mv-nav-links">
      {nav_links}
    </div>
  </nav>
  <div class="wrap">
"""


def hero_logo():
    """Big logo, same treatment as index.html, for the top of every other page."""
    return """    <div style="text-align:center;margin-bottom:36px;">
      <img src="logo-trim.png" alt="MEGAVISION" class="mv-glow" style="width:100%;max-width:640px;height:auto;">
    </div>
"""


def foot():
    updated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return f"""  </div>
  <footer class="mv-footer">MEGAVISION &middot; Mega League Archive &middot; <a href="style-guide.html">Style Guide</a> &middot; Updated {updated_at}</footer>
  <script>
    function mvShowTab(btn, panelId) {{
      var tabs = btn.parentElement.querySelectorAll('.mv-tab');
      var panels = btn.closest('.mv-card').querySelectorAll('.mv-tab-panel');
      tabs.forEach(function(t) {{ t.classList.remove('active'); }});
      panels.forEach(function(p) {{ p.classList.remove('active'); }});
      btn.classList.add('active');
      document.getElementById(panelId).classList.add('active');
    }}

    // Any <table class="mv-sortable"> gets click-to-sort headers for free.
    // Mark a <th> with data-sort-type="num" or "text" to make it sortable;
    // give a <td> a data-sort="..." attribute to sort by a value other than
    // its displayed text. Only <tbody> rows are reordered -- put running
    // totals in <tfoot> so they stay pinned.
    function mvMakeSortable(table) {{
      var tbody = table.querySelector('tbody');
      if (!tbody) return;
      table.querySelectorAll('th').forEach(function(th, idx) {{
        if (!th.dataset.sortType) return;
        th.style.cursor = 'pointer';
        var dir = 1;
        th.addEventListener('click', function() {{
          var rows = Array.from(tbody.querySelectorAll('tr'));
          var type = th.dataset.sortType;
          rows.sort(function(a, b) {{
            var ac = a.children[idx], bc = b.children[idx];
            if (!ac || !bc) return 0;
            var av = ac.dataset.sort !== undefined ? ac.dataset.sort : ac.textContent.trim();
            var bv = bc.dataset.sort !== undefined ? bc.dataset.sort : bc.textContent.trim();
            if (type === 'num') {{
              av = parseFloat(av); bv = parseFloat(bv);
              if (isNaN(av)) av = -Infinity;
              if (isNaN(bv)) bv = -Infinity;
            }}
            if (av < bv) return -1 * dir;
            if (av > bv) return 1 * dir;
            return 0;
          }});
          dir *= -1;
          rows.forEach(function(r) {{ tbody.appendChild(r); }});
        }});
      }});
    }}
    document.querySelectorAll('table.mv-sortable').forEach(mvMakeSortable);
  </script>
</body>
</html>
"""


# kept for any old callers; prefer foot() so the timestamp is current
FOOT = foot()
